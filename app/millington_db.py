# db.py — database connection and all query functions
# =============================================================================
# All Supabase calls live here. No other file talks to the database directly.
# This means if the database schema changes, there is only one file to update.
#
# Connection:
#   - Locally: reads SUPABASE_URL and SUPABASE_KEY from .env
#   - On Streamlit Cloud: reads from st.secrets
#
# Every function returns plain Python dicts or lists — never raw Supabase
# response objects. This keeps the rest of the app simple.
# =============================================================================

import os
import re
from datetime import date
import streamlit as st
from supabase import create_client, Client
from dotenv import load_dotenv
from rapidfuzz import process, fuzz

load_dotenv()

# SKU pattern embedded in Holded product names e.g. "Cookie Box - CO-03-DC-GW"
_SKU_RE = re.compile(r'\b([A-Z]{2}-\d{2}-[A-Z]{2}-[A-Z]{2,4}(?:-[A-Z]{2})?)\b')
 
# Spanish month name → month number
_MONTHS_ES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}

# Helpers

def _normalise_name(name: str) -> str:
    """Strip leading/trailing whitespace and collapse internal spaces."""
    return re.sub(r'\s+', ' ', name).strip()

def find_similar_names(name: str, existing_names: list[str], 
                        threshold: int = 85) -> list[tuple[str, int]]:
    """
    Return existing names that are suspiciously similar to the proposed name.
    Uses token sort ratio which handles word order differences, e.g.
    'Chocolate Negro 70%' vs '70% Chocolate Negro' would still match.
    Returns a list of (name, score) tuples above the threshold,
    sorted by score descending.
    """
    name_normalised = _normalise_name(name)
    if not name_normalised:
        return []
    results = process.extract(
        name_normalised,
        existing_names,
        scorer=fuzz.token_sort_ratio,
        limit=3,
    )
    # Filter out exact matches (the name itself if editing) and low scores
    return [
        (match, score)
        for match, score, _ in results
        if score >= threshold and match.lower() != name_normalised.lower()
    ]

def get_current_prices(cake_code: str) -> list[dict]:
    """
    Return all current_prices rows for a given cake code prefix.
    e.g. cake_code='LP' returns all LP-* SKUs across all channels.
    Returns [] gracefully if the table doesn't exist yet.
    """
    sb = get_client()
    try:
        # Use range filter instead of ilike to avoid Cloudflare edge issues
        # with % wildcards.  "LP-" … "LP." captures all LP-* SKU codes because
        # "." (ASCII 46) is the character immediately above "-" (ASCII 45).
        result = (
            sb.table("current_prices")
            .select("*")
            .gte("sku_code", f"{cake_code}-")
            .lt("sku_code",  f"{cake_code}.")
            .order("sku_code")
            .execute()
        )
        return result.data or []
    except Exception as e:
        import streamlit as _st
        _st.session_state["_current_prices_error"] = str(e)
        return []

# =============================================================================
# Recipe weight estimation
# =============================================================================

# Known weights for unit-based ingredients (grams per unit)
_UNIT_WEIGHTS_G = {
    "huevos":    50.0,   # medium egg, net edible weight
    "manzanas":  150.0,  # medium apple
    "limones":   65.0,   # medium-large lemon, juice + zest combined yield
    "limas":     45.0,   # medium-large lime, juice + zest combined yield
    "naranja":   80.0,   # medium orange, juice + zest combined yield
}

# How unit-count recipe amounts should be converted for purchasing display.
# Tuple of (purchase_unit, factor): recipe_count × factor = purchase_amount.
_UNIT_PURCHASE = {
    "huevos":   ("docenas", 1 / 12),   # eggs → dozens
    "manzanas": ("kg",      0.150),    # apples → kg
    "limones":  ("kg",      0.065),    # lemons → kg
    "limas":    ("kg",      0.045),    # limes → kg
    "naranja":  ("kg",      0.080),    # oranges → kg
}

# Unit ingredients to silently ignore (only part used, weight negligible,
# or weight not meaningful for costing)
_UNIT_IGNORE = {
    "vainilla rama",
    "canela en rama",
}

def _to_label_grams(name: str, amount: float) -> float:
    """
    Convert a raw ingredient amount to grams for label weight ordering.
    Unit-based ingredients (eggs, lemons etc.) are multiplied by their
    standard gram equivalent from _UNIT_WEIGHTS_G.
    Ingredients in _UNIT_IGNORE return 0 (excluded from weight ordering).
    All others are assumed to already be in grams.
    """
    if amount is None:
        return 0.0
    name_lower = (name or "").lower()
    for key, weight in _UNIT_WEIGHTS_G.items():
        if key in name_lower:
            return amount * weight
    if any(key in name_lower for key in _UNIT_IGNORE):
        return 0.0
    return float(amount)

def estimate_recipe_weight(lines: list[dict]) -> dict:
    """
    Estimate the finished weight of a recipe in grams by summing
    ingredient amounts.

    Recipe amounts are ALWAYS in grams or units — pack_unit on the
    ingredient record describes the purchase pack and is irrelevant here.

    Rules:
      - All numeric amounts: add directly as grams
      - Unit ingredients with known weight (eggs, apples): multiply
      - Unit ingredients in _UNIT_IGNORE: skip silently
      - All other unit ingredients: exclude and flag
    """
    total_g  = 0.0
    excluded = []
    notes    = []

    for line in lines:
        ing_name = (line.get("ingredient_name") or "").strip()
        amount   = float(line.get("amount") or 0)

        if not ing_name or amount <= 0:
            continue

        name_lower = ing_name.lower()

        # Check if this is a known unit ingredient
        matched_weight = next(
            (w for key, w in _UNIT_WEIGHTS_G.items()
             if key in name_lower),
            None
        )

        if matched_weight is not None:
            # Unit ingredient with known weight — e.g. eggs
            grams = amount * matched_weight
            total_g += grams
            notes.append(
                f"{ing_name}: {amount:.0f} × "
                f"{matched_weight:.0f}g = {grams:.0f}g"
            )
        elif any(key in name_lower for key in _UNIT_IGNORE):
            # Known unit ingredients to ignore (lemons, limes etc.)
            pass
        elif amount < 20:
            # Small amounts likely to be unit-based (e.g. 1 vanilla pod,
            # 4 gelatine sheets) — flag rather than add raw
            excluded.append(f"{ing_name} ({amount:.0f})")
        else:
            # Treat as grams directly
            total_g += amount

    return {
        "weight_g": round(total_g, 1),
        "excluded": excluded,
        "notes":    notes,
    }

# -----------------------------------------------------------------------------
# Connection
# -----------------------------------------------------------------------------

@st.cache_resource
def get_client() -> Client:
    """
    Create and cache a single Supabase client for the app's lifetime.
    st.cache_resource means this runs once and reuses the connection.
    """
    # Try Streamlit secrets first (production), fall back to .env (local dev)
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
    except (KeyError, FileNotFoundError):
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_KEY")

    if not url or not key:
        st.error(
            "Database credentials not found. "
            "Add SUPABASE_URL and SUPABASE_KEY to your .env file."
        )
        st.stop()

    return create_client(url, key)


# -----------------------------------------------------------------------------
# Ingredient stock
# -----------------------------------------------------------------------------

def get_ingredient_stock() -> dict[str, float]:
    """Return current stock as {ingredient_id: amount}."""
    sb = get_client()
    result = sb.table("ingredient_stock").select("ingredient_id, amount").execute()
    return {row["ingredient_id"]: float(row["amount"]) for row in (result.data or [])}


def upsert_ingredient_stock(ingredient_id: str, amount: float) -> None:
    """Set current stock level for one ingredient."""
    sb = get_client()
    sb.table("ingredient_stock").upsert({
        "ingredient_id": ingredient_id,
        "amount":        amount,
        "updated_at":    "now()",
    }).execute()


def get_recipe_ingredients_for_shopping() -> list[dict]:
    """
    Return all recipe ingredient lines with ingredient name/pack_unit
    and the recipe's ref_batch_size. Used by the shopping list screen.

    Returns list of dicts:
      recipe_id, ref_batch_size, ingredient_id, ingredient_name, pack_unit, amount
    """
    sb = get_client()

    # Fetch ingredient lines with ingredient details
    lines_result = (
        sb.table("recipe_ingredient_lines")
        .select("recipe_id, amount, ingredients(id, name, pack_unit)")
        .execute()
    )

    # Fetch recipe ref_batch_size separately (reverse join not always available)
    recipes_result = (
        sb.table("recipes")
        .select("id, ref_batch_size")
        .execute()
    )
    batch_by_recipe = {
        r["id"]: float(r.get("ref_batch_size") or 1)
        for r in (recipes_result.data or [])
    }

    rows = []
    for row in lines_result.data or []:
        ing = row.pop("ingredients", None) or {}
        recipe_id = row.get("recipe_id", "")
        # Recipe amounts are always in base units (g, ml, or units).
        # pack_unit describes the purchase pack (e.g. "kg" bag) — normalise
        # it back to the base unit so display is consistent.
        raw_unit  = (ing.get("pack_unit") or "g").lower().strip()
        base_unit = {"kg": "g", "l": "ml", "litre": "ml", "litro": "ml"}.get(raw_unit, raw_unit)

        # Known unit-count ingredients (lemons, eggs etc.): convert recipe
        # count to a practical purchase unit (kg for fruit, dozens for eggs).
        ing_name_lower = (ing.get("name") or "").lower()
        purchase_key   = next(
            (k for k in _UNIT_PURCHASE if k in ing_name_lower), None
        )
        is_unit_ignore = any(key in ing_name_lower for key in _UNIT_IGNORE)
        raw_amount     = float(row.get("amount") or 0)

        if purchase_key:
            p_unit, p_factor = _UNIT_PURCHASE[purchase_key]
            base_unit  = p_unit
            raw_amount = raw_amount * p_factor
        elif is_unit_ignore:
            base_unit = "units"

        rows.append({
            "recipe_id":       recipe_id,
            "ref_batch_size":  batch_by_recipe.get(recipe_id, 1.0),
            "ingredient_id":   ing.get("id",       ""),
            "ingredient_name": ing.get("name",      ""),
            "pack_unit":       base_unit,
            "unit_weight_g":   None,   # no longer needed — conversion already applied
            "amount":          raw_amount,
        })
    return rows


# -----------------------------------------------------------------------------
# Ingredients
# -----------------------------------------------------------------------------

def get_ingredients() -> list[dict]:
    sb = get_client()
    result = sb.table("ingredients").select("*").order("name").execute()
    ingredients = result.data or []

    # Compute derived costs for ingredients that reference a base ingredient.
    # Formula: cost_per_unit = (base.cost_per_unit × cost_fraction) / yield_g_per_base_unit
    id_map = {i["id"]: i for i in ingredients}
    for ing in ingredients:
        base_id = ing.get("base_ingredient_id")
        if not base_id:
            continue
        base = id_map.get(base_id)
        if not base:
            continue
        base_cost = base.get("cost_per_unit") or 0.0
        yield_g   = ing.get("yield_g_per_base_unit") or 1.0
        fraction  = ing.get("cost_fraction") or 1.0
        ing["cost_per_unit"] = round((base_cost * fraction) / yield_g, 6) if yield_g else None
        # Inherit allergen category from base if not explicitly set
        if not ing.get("category_id"):
            ing["category_id"] = base.get("category_id")

    return ingredients

def get_ingredient_categories() -> list[dict]:
    sb = get_client()
    result = sb.table("ingredient_categories").select("*").order("label_name_es").execute()
    return result.data or []


def save_ingredient(record: dict) -> dict:
    """Insert or update an ingredient. Computes cost_per_unit before saving.
    For derived ingredients (base_ingredient_id set), cost_per_unit is computed
    at query time from the base; pack price fields are not required.
    Category is inherited from the base ingredient for allergen resolution.
    """
    sb = get_client()
    record["name"]       = _normalise_name(record.get("name", ""))
    record["updated_at"] = "now()"
    if record.get("base_ingredient_id"):
        # Derived ingredient — inherit category from base, don't overwrite cost
        base_rows = (
            sb.table("ingredients")
              .select("category_id")
              .eq("id", record["base_ingredient_id"])
              .execute()
              .data or []
        )
        if base_rows and not record.get("category_id"):
            record["category_id"] = base_rows[0].get("category_id")
        # cost_per_unit is computed at read time; store None to avoid stale values
        record["cost_per_unit"] = None
    else:
        record = _compute_ingredient_cost(record)
    if record.get("id"):
        sb.table("ingredients").update(record).eq("id", record["id"]).execute()
        result = sb.table("ingredients").select("*").eq("id", record["id"]).execute()
    else:
        result = sb.table("ingredients").insert(record).execute()
    return result.data[0] if result.data else {}


def delete_ingredient(ingredient_id: str) -> None:
    sb = get_client()
    sb.table("ingredients").delete().eq("id", ingredient_id).execute()


# Conversion factors to base units (g for weight, ml for volume)
_UNIT_TO_BASE = {
    "g":     1.0,
    "kg":    1000.0,
    "ml":    1.0,
    "l":     1000.0,
    "units": 1.0,   # units stay as units — recipe amounts are also in units
}

def _compute_ingredient_cost(record: dict) -> dict:
    """
    Compute cost_per_unit from pack_price_ex_vat and pack_size.
    Always normalises to cost per base unit:
      - weight ingredients → cost per gram
      - volume ingredients → cost per ml
      - unit ingredients   → cost per unit
    This ensures recipe amounts (always in grams, ml or units)
    multiply correctly regardless of how the pack size was entered.
    """
    try:
        price     = float(record.get("pack_price_ex_vat") or 0)
        size      = float(record.get("pack_size") or 0)
        unit      = record.get("pack_unit") or "g"
        # Convert pack size to base units before dividing
        factor    = _UNIT_TO_BASE.get(unit, 1.0)
        base_size = size * factor
        record["cost_per_unit"] = round(price / base_size, 6) if base_size > 0 else None
    except (TypeError, ValueError):
        record["cost_per_unit"] = None
    return record

def save_ingredient_allergens(record: dict) -> None:
    """Save allergen and ficha fields for an ingredient."""
    sb = get_client()
    allowed = {
        k: v for k, v in record.items()
        if k.startswith("allergen_")
        or k in (
            "id", "category_id", "allergen_override", "is_sub_recipe",
            "label_name_es",
            "label_name_es_2", "label_name_es_2_pct",
            "label_name_es_3", "label_name_es_3_pct",
        )
    }
    allowed["updated_at"] = "now()"
    sb.table("ingredients").update(allowed).eq("id", allowed["id"]).execute()

# -----------------------------------------------------------------------------
# Consumables
# -----------------------------------------------------------------------------

def get_consumables() -> list[dict]:
    sb = get_client()
    result = sb.table("consumables").select("*").order("name").execute()
    return result.data or []


def save_consumable(record: dict) -> dict:
    """Insert or update a consumable. Computes cost_per_unit before saving."""
    sb = get_client()
    record["name"] = _normalise_name(record.get("name", ""))
    record = _compute_consumable_cost(record)
    if record.get("id"):
        sb.table("consumables").update(record).eq("id", record["id"]).execute()
        result = sb.table("consumables").select("*").eq("id", record["id"]).execute()
    else:
        result = sb.table("consumables").insert(record).execute()
    return result.data[0] if result.data else {}


def delete_consumable(consumable_id: str) -> None:
    sb = get_client()
    sb.table("consumables").delete().eq("id", consumable_id).execute()


def _compute_consumable_cost(record: dict) -> dict:
    """Compute cost_per_unit from pack_price_ex_vat and pack_quantity."""
    try:
        price = float(record.get("pack_price_ex_vat") or 0)
        qty   = float(record.get("pack_quantity") or 0)
        record["cost_per_unit"] = round(price / qty, 6) if qty > 0 else None
    except (TypeError, ValueError):
        record["cost_per_unit"] = None
    return record


# -----------------------------------------------------------------------------
# Recipes
# -----------------------------------------------------------------------------

def get_recipes(
    include_sub_recipes:  bool = False,
    include_deprecated:   bool = False,
) -> list[dict]:
    """Return recipes joined with cake_codes so screens can read the
    cake code string without a second DB call."""
    sb = get_client()
    q  = (sb.table("recipes")
            .select("*, cake_codes(code, name)")
            .order("name"))
    if not include_sub_recipes:
        q = q.eq("is_sub_recipe", False)
    if not include_deprecated:
        q = q.eq("deprecated", False)
    return q.execute().data or []


def get_component_recipes() -> list[dict]:
    """Return all component (sub) recipes — for ingredient pickers and production log."""
    sb = get_client()
    return (
        sb.table("recipes")
          .select("id, name, labour_per_kg, ref_weight_kg")
          .eq("is_sub_recipe", True)
          .eq("deprecated", False)
          .order("name")
          .execute()
          .data or []
    )


def get_recipe(recipe_id: str) -> dict:
    sb = get_client()
    result = sb.table("recipes").select("*").eq("id", recipe_id).execute()
    return result.data[0] if result.data else {}


def save_recipe(record: dict) -> dict:
    sb = get_client()
    if record.get("id"):
        sb.table("recipes").update(record).eq("id", record["id"]).execute()
        result = sb.table("recipes").select("*").eq("id", record["id"]).execute()
    else:
        insert_record = {k: v for k, v in record.items() if k != "id"}
        result = sb.table("recipes").insert(insert_record).execute()
    return result.data[0] if result.data else {}


def delete_recipe(recipe_id: str) -> None:
    sb = get_client()
    # ingredient lines cascade-delete automatically (ON DELETE CASCADE)
    sb.table("recipes").delete().eq("id", recipe_id).execute()


def duplicate_recipe(recipe_id: str, new_name: str | None = None) -> dict:
    """
    Copy a recipe (fields + ingredient lines + PCC steps) to a new record.
    The copy has no cake_code_id, deprecated=False, and a name suffixed
    with ' [copia]' unless new_name is provided.
    Variants are NOT copied — they remain on the original.
    Returns the new recipe dict.
    """
    sb     = get_client()
    source = get_recipe(recipe_id)
    if not source:
        raise ValueError(f"Recipe {recipe_id} not found")

    EXCLUDE = {"id", "created_at", "updated_at"}
    new_rec = {k: v for k, v in source.items() if k not in EXCLUDE}
    new_rec["name"]          = new_name or f"{source['name']} [copia]"
    new_rec["cake_code_id"]  = None   # unassigned — keeps it out of active screens
    new_rec["deprecated"]    = False

    saved = sb.table("recipes").insert(new_rec).execute().data[0]
    new_id = saved["id"]

    # Copy ingredient lines
    lines = (
        sb.table("recipe_ingredient_lines")
          .select("ingredient_id, component_recipe_id, amount, sort_order")
          .eq("recipe_id", recipe_id)
          .order("sort_order")
          .execute()
          .data or []
    )
    if lines:
        for line in lines:
            line["recipe_id"] = new_id
        sb.table("recipe_ingredient_lines").insert(lines).execute()

    # Copy PCC steps
    steps = (
        sb.table("recipe_pcc_steps")
          .select("step_name, target_temp_c, target_time_min, critical_limit_temp_c, sort_order")
          .eq("recipe_id", recipe_id)
          .order("sort_order")
          .execute()
          .data or []
    )
    if steps:
        for step in steps:
            step["recipe_id"] = new_id
        sb.table("recipe_pcc_steps").insert(steps).execute()

    return saved


# -----------------------------------------------------------------------------
# Recipe ingredient lines
# -----------------------------------------------------------------------------

def get_recipe_lines(recipe_id: str) -> list[dict]:
    """
    Return ingredient lines for a recipe, joined with ingredient names
    and costs so the UI does not need to do extra lookups.
    Handles both raw ingredient lines and component recipe lines.
    """
    sb = get_client()
    result = (
        sb.table("recipe_ingredient_lines")
        .select("*, ingredients(name, cost_per_unit, pack_unit)")
        .eq("recipe_id", recipe_id)
        .order("sort_order")
        .execute()
    )
    # Batch-fetch component recipe details for component lines
    comp_ids = [
        row["component_recipe_id"]
        for row in (result.data or [])
        if row.get("component_recipe_id")
    ]
    comp_map: dict = {}
    if comp_ids:
        comp_rows = (
            sb.table("recipes")
              .select("id, name, labour_per_kg, ref_weight_kg")
              .in_("id", comp_ids)
              .execute()
              .data or []
        )
        comp_map = {r["id"]: r for r in comp_rows}

    lines = []
    for row in result.data or []:
        ing     = row.pop("ingredients", None) or {}
        comp_id = row.get("component_recipe_id")
        if comp_id:
            comp = comp_map.get(comp_id, {})
            row["ingredient_name"]          = comp.get("name", "")
            row["ingredient_cost_per_unit"] = None
            row["ingredient_unit"]          = "g"
            row["is_component_line"]        = True
            row["component_labour_per_kg"]  = comp.get("labour_per_kg")
            row["component_ref_weight_kg"]  = comp.get("ref_weight_kg")
        else:
            row["ingredient_name"]          = ing.get("name", "")
            row["ingredient_cost_per_unit"] = ing.get("cost_per_unit")
            row["ingredient_unit"]          = ing.get("pack_unit", "g")
            row["is_component_line"]        = False
        lines.append(row)
    return lines


def save_recipe_line(record: dict) -> dict:
    sb = get_client()
    if record.get("id"):
        sb.table("recipe_ingredient_lines").update(record).eq("id", record["id"]).execute()
        result = sb.table("recipe_ingredient_lines").select("*").eq("id", record["id"]).execute()
    else:
        result = sb.table("recipe_ingredient_lines").insert(record).execute()
    return result.data[0] if result.data else {}


def delete_recipe_line(line_id: str) -> None:
    sb = get_client()
    sb.table("recipe_ingredient_lines").delete().eq("id", line_id).execute()


def replace_recipe_lines(recipe_id: str, lines: list[dict]) -> None:
    """
    Replace all ingredient lines for a recipe in a single operation.
    Handles both raw ingredient lines (ingredient_id) and component
    recipe lines (component_recipe_id).
    """
    sb = get_client()
    sb.table("recipe_ingredient_lines").delete().eq("recipe_id", recipe_id).execute()
    if lines:
        rows = []
        for i, line in enumerate(lines):
            row: dict = {
                "recipe_id":  recipe_id,
                "sort_order": i,
                "amount":     line.get("amount"),
            }
            if line.get("component_recipe_id"):
                row["component_recipe_id"] = line["component_recipe_id"]
                row["ingredient_id"]       = None
            else:
                row["ingredient_id"]       = line.get("ingredient_id")
                row["component_recipe_id"] = None
            rows.append(row)
        sb.table("recipe_ingredient_lines").insert(rows).execute()
        
def get_all_variants() -> list[dict]:
    """Fetch all product variants in one query — used for sidebar counts."""
    sb = get_client()
    result = (
        sb.table("product_variants")
        .select("id, recipe_id, format, label_approved, channel")
        .execute()
    )
    return result.data or []

def get_variants_for_recipe(recipe_id: str) -> list[dict]:
    sb = get_client()
    result = (
        sb.table("product_variants")
        .select("*")
        .eq("recipe_id", recipe_id)
        .order("format")
        .execute()
    )
    return result.data or []


def save_variant(record: dict) -> dict:
    sb = get_client()
    record["updated_at"] = "now()"
    record.pop("sku_code", None)
    # Timestamp price fields when they are updated
    if "ws_price_ex_vat" in record:
        record["ws_price_updated_at"] = "now()"
    if "rt_price_inc_vat" in record:
        record["rt_price_updated_at"] = "now()"
    if record.get("id"):
        sb.table("product_variants").update(record).eq(
            "id", record["id"]
        ).execute()
        result = sb.table("product_variants").select("*").eq(
            "id", record["id"]
        ).execute()
    else:
        result = sb.table("product_variants").insert(record).execute()
    return result.data[0] if result.data else {}


def delete_variant(variant_id: str) -> None:
    sb = get_client()
    sb.table("product_variants").delete().eq("id", variant_id).execute()


def get_size_code_definitions() -> list[dict]:
    """All size code definitions ordered by tier then sort_order."""
    return (
        get_client()
        .table("size_code_definitions")
        .select("*")
        .order("sort_order")
        .execute()
        .data or []
    )


def save_size_code_definition(record: dict) -> None:
    sb  = get_client()
    exists = (
        sb.table("size_code_definitions")
        .select("code")
        .eq("code", record["code"])
        .execute()
        .data
    )
    if exists:
        sb.table("size_code_definitions").update(record).eq("code", record["code"]).execute()
    else:
        sb.table("size_code_definitions").insert(record).execute()


def delete_size_code_definition(code: str) -> None:
    get_client().table("size_code_definitions").delete().eq("code", code).execute()


def get_sku_weight_map() -> dict[str, float]:
    """
    Returns {full_sku_code: ref_weight_g} for all variants that have both.
    Used by screen_kpis._scale_factor to interpolate costs for non-standard
    size codes (CA, LI, etc.) using the weight stored in the variant record.
    """
    sb = get_client()
    result = (
        sb.table("product_variants")
        .select("sku_ws, sku_gw, ref_weight_g")
        .execute()
    )
    out: dict[str, float] = {}
    for row in (result.data or []):
        w = row.get("ref_weight_g")
        if not w:
            continue
        for field in ("sku_ws", "sku_gw"):
            code = row.get(field)
            if code:
                out[code] = float(w)
    return out


def get_all_variants_full() -> list[dict]:
    """Fetch all variants with working, approved price fields and size info."""
    sb = get_client()
    result = (
        sb.table("product_variants")
        .select(
            "id, recipe_id, format, channel, size_description, "
            "ref_diameter_cm, ref_height_cm, "
            "ws_price_ex_vat, ws_price_approved, ws_price_approved_at, "
            "rt_price_inc_vat, rt_price_approved, rt_price_approved_at, "
            "label_approved"
        )
        .execute()
    )
    return result.data or []

def get_ingredient_lines_all() -> list[dict]:
    """
    Return every recipe_ingredient_lines row joined with its ingredient's
    name, cost_per_unit and pack_unit.
 
    Used by screen_kpis.py to compute estimated ingredient spend from
    Holded sales data without N+1 queries.
 
    Returns a list of dicts with keys:
      recipe_id, amount, ingredient_id, ingredient_name,
      cost_per_unit, pack_unit
    """
    sb     = get_client()
    result = (
        sb.table("recipe_ingredient_lines")
        .select(
            "recipe_id, amount, "
            "ingredients(id, name, cost_per_unit, pack_unit)"
        )
        .execute()
    )
    rows = []
    for row in result.data or []:
        ing = row.pop("ingredients", None) or {}
        row["ingredient_id"]   = ing.get("id",             "")
        row["ingredient_name"] = ing.get("name",           "Unknown")
        row["cost_per_unit"]   = ing.get("cost_per_unit")
        row["pack_unit"]       = ing.get("pack_unit",      "")
        rows.append(row)
    return rows

# =============================================================================
# ALLERGEN DECLARATION GENERATOR
# Add these functions to millington_db.py
# =============================================================================
#
# These functions build the legal allergen declaration for a recipe ficha.
#
# EU 1169/2011 / RD 126/2015 terminology — Spanish legal text.
#
# Logic:
#   1. Fetch ingredient lines for the recipe
#   2. For each line, resolve effective allergen values:
#        - If ingredient.allergen_override = TRUE → use ingredient allergen fields
#        - Otherwise → use ingredient.category allergen fields
#   3. If ingredient.is_sub_recipe = TRUE → recurse into matching recipe
#   4. Union all allergen values (max wins: 2 > 1 > 0)
#   5. Add kitchen_may_contain from the recipe
#   6. Return structured Contiene / Puede contener lists
#
# Also generates a draft ingredient label text ordered by weight descending.
# =============================================================================


# ── Allergen field names and Spanish legal display text ───────────────────────
ALLERGEN_FIELDS = [
    "allergen_gluten",
    "allergen_crustacean",
    "allergen_egg",
    "allergen_fish",
    "allergen_peanut",
    "allergen_soy",
    "allergen_milk",
    "allergen_nuts",
    "allergen_celery",
    "allergen_mustard",
    "allergen_sesame",
    "allergen_sulphites",
    "allergen_lupin",
    "allergen_mollusc",
]

ALLERGEN_DISPLAY_ES = {
    "allergen_gluten":     "cereales con gluten y sus derivados",
    "allergen_crustacean": "crustáceos y sus derivados",
    "allergen_egg":        "huevo y derivados",
    "allergen_fish":       "pescado y sus derivados",
    "allergen_peanut":     "cacahuetes y sus derivados",
    "allergen_soy":        "soja y sus derivados",
    "allergen_milk":       "leche y derivados lácteos",
    "allergen_nuts":       "frutos de cáscara y sus derivados",
    "allergen_celery":     "apio y sus derivados",
    "allergen_mustard":    "mostaza y sus derivados",
    "allergen_sesame":     "granos de sésamo y sus derivados",
    "allergen_sulphites":  "dióxido de azufre y sulfitos",
    "allergen_lupin":      "altramuces y sus derivados",
    "allergen_mollusc":    "moluscos y sus derivados",
}


def _get_recipe_lines_with_allergens(recipe_id: str) -> list[dict]:
    """
    Fetch ingredient lines for a recipe with full ingredient data:
    allergen fields, category allergen fields, is_sub_recipe, label_name_es.
    Also handles component recipe lines (component_recipe_id set, ingredient_id null).
    """
    sb = get_client()

    result = (
        sb.table("recipe_ingredient_lines")
        .select(
            "amount, sort_order, component_recipe_id, "
            "ingredients("
            "  id, name, label_name_es, label_name_es_2, label_name_es_2_pct, "
            "  label_name_es_3, label_name_es_3_pct, "
            "  is_sub_recipe, allergen_override, allergen_notes, "
            + ", ".join(ALLERGEN_FIELDS) + ", "
            "  ingredient_categories(id, label_name_es, " +
            ", ".join(ALLERGEN_FIELDS) + ")"
            ")"
        )
        .eq("recipe_id", recipe_id)
        .order("sort_order")
        .execute()
    )

    # Batch-fetch component recipe names
    comp_ids = [
        row["component_recipe_id"]
        for row in (result.data or [])
        if row.get("component_recipe_id")
    ]
    comp_name_map: dict = {}
    if comp_ids:
        comp_rows = (
            sb.table("recipes")
              .select("id, name")
              .in_("id", comp_ids)
              .execute()
              .data or []
        )
        comp_name_map = {r["id"]: r["name"] for r in comp_rows}

    lines = []
    for row in result.data or []:
        ing     = row.pop("ingredients", None) or {}
        cat     = ing.pop("ingredient_categories", None) or {}
        comp_id = row.get("component_recipe_id")

        if comp_id:
            # Component recipe line — allergens resolved by direct recursion
            comp_name = comp_name_map.get(comp_id, "")
            entry = {
                "amount":               float(row.get("amount") or 0),
                "sort_order":           row.get("sort_order", 0),
                "ingredient_id":        None,
                "ingredient_name":      comp_name,
                "label_name_es":        comp_name,
                "label_name_es_2":      None,
                "label_name_es_2_pct":  None,
                "label_name_es_3":      None,
                "label_name_es_3_pct":  None,
                "is_sub_recipe":        True,
                "is_component_line":    True,
                "component_recipe_id":  comp_id,
                "allergen_override":    False,
                "allergen_notes":       None,
                "category_label":       "",
                "category":             {},
                "ingredient":           {},
            }
        else:
            ing_label = (
                ing.get("label_name_es")
                or cat.get("label_name_es")
                or ing.get("name", "")
            )
            entry = {
                "amount":               float(row.get("amount") or 0),
                "sort_order":           row.get("sort_order", 0),
                "ingredient_id":        ing.get("id"),
                "ingredient_name":      ing.get("name", ""),
                "label_name_es":        ing_label,
                "label_name_es_2":      ing.get("label_name_es_2"),
                "label_name_es_2_pct":  ing.get("label_name_es_2_pct"),
                "label_name_es_3":      ing.get("label_name_es_3"),
                "label_name_es_3_pct":  ing.get("label_name_es_3_pct"),
                "is_sub_recipe":        bool(ing.get("is_sub_recipe")),
                "is_component_line":    False,
                "component_recipe_id":  None,
                "allergen_override":    bool(ing.get("allergen_override")),
                "allergen_notes":       ing.get("allergen_notes"),
                "category_label":       cat.get("label_name_es", ""),
                "category":             cat,
                "ingredient":           ing,
            }
        lines.append(entry)

    return lines


def _find_recipe_by_ingredient_name(name: str) -> dict:
    """
    Find a recipe whose name loosely matches an ingredient marked as sub_recipe.
    Used to expand sub-recipes during allergen calculation.
    """
    sb = get_client()
    result = (
        sb.table("recipes")
        .select("id, name, kitchen_may_contain")
        .ilike("name", f"%{name}%")
        .limit(1)
        .execute()
    )
    return result.data[0] if result.data else {}


def _effective_allergens(line: dict) -> dict:
    """
    Return the effective allergen values for one ingredient line.
    Uses ingredient fields if allergen_override=True,
    otherwise uses category fields.
    Returns {field: value} for all 14 allergens.
    """
    if line["allergen_override"]:
        source = line["ingredient"]
    else:
        source = line["category"]

    return {
        field: int(source.get(field) or 0)
        for field in ALLERGEN_FIELDS
    }


def _union_allergens(
    accumulated: dict[str, int],
    new_values: dict[str, int]
) -> dict[str, int]:
    """
    Merge two allergen dicts — highest value wins.
    0=no, 1=contiene, 2=puede contener.
    Note: 1 (Contiene) beats 2 (Puede contener) since definite presence
    is more serious than possible presence.
    Special merge rule:
      - If accumulated=0 and new=2 → result=2
      - If accumulated=2 and new=1 → result=1 (Contiene overrides Puede)
      - If accumulated=1 and new=2 → result=1 (keep Contiene)
      - Otherwise max wins
    """
    result = {}
    for field in ALLERGEN_FIELDS:
        a = accumulated.get(field, 0)
        n = new_values.get(field, 0)
        if a == 0:
            result[field] = n
        elif a == 1:
            result[field] = 1  # Contiene always wins
        elif a == 2:
            result[field] = 1 if n == 1 else 2
        else:
            result[field] = max(a, n)
    return result


def get_allergen_declaration(
    recipe_id: str,
    depth: int = 0,
    _visited: set | None = None
) -> dict:
    """
    Build the full allergen declaration for a recipe.

    Returns:
    {
        "contiene":          [list of Spanish legal allergen strings],
        "puede_contener":    [list of Spanish legal allergen strings],
        "accumulated":       {field: 0|1|2} raw values,
        "warnings":          [list of warning strings],
        "ingredient_names":  [list of (label_name_es, amount_g) for label text],
    }

    Args:
        recipe_id: UUID of the recipe to process
        depth:     recursion depth (max 5 to prevent infinite loops)
        _visited:  set of recipe_ids already visited (loop detection)
    """
    if _visited is None:
        _visited = set()

    if depth > 5:
        return {
            "contiene": [], "puede_contener": [], "accumulated": {},
            "warnings": ["⚠️ Profundidad máxima de recursión alcanzada"],
            "ingredient_names": [],
        }

    if recipe_id in _visited:
        return {
            "contiene": [], "puede_contener": [], "accumulated": {},
            "warnings": [f"⚠️ Referencia circular detectada en receta {recipe_id}"],
            "ingredient_names": [],
        }

    _visited.add(recipe_id)

    # Fetch recipe for kitchen_may_contain
    recipe  = get_recipe(recipe_id)
    kitchen = recipe.get("kitchen_may_contain") or ""

    # Fetch ingredient lines with allergen data
    lines    = _get_recipe_lines_with_allergens(recipe_id)
    warnings = []

    # Accumulate allergen values across all ingredients
    accumulated   = {f: 0 for f in ALLERGEN_FIELDS}
    ing_for_label = []  # [(label_name_es, amount_g), ...]

    for line in lines:
        name   = line["ingredient_name"]
        amount = line["amount"]

        if line["is_sub_recipe"]:
            # Resolve component recipe: use direct FK if available, else fuzzy match
            comp_id    = line.get("component_recipe_id")
            sub_recipe = get_recipe(comp_id) if comp_id else _find_recipe_by_ingredient_name(name)
            if sub_recipe:
                sub_result = get_allergen_declaration(
                    sub_recipe["id"],
                    depth=depth + 1,
                    _visited=_visited
                )
                accumulated = _union_allergens(
                    accumulated, sub_result["accumulated"]
                )
                warnings.extend(sub_result["warnings"])
                # Scale sub-recipe ingredient amounts correctly.
                # sub_ing[1] is already an absolute gram quantity within the
                # sub-recipe's own reference batch. Scale by
                # (amount_used / sub_recipe_total_g) — not by amount directly,
                # which would give grams² (the bug being fixed here).
                sub_total_g = sum(
                    _to_label_grams(s[0], s[1])
                    for s in sub_result["ingredient_names"]
                ) or amount
                sub_scale = amount / sub_total_g if sub_total_g else 1.0
                for sub_ing in sub_result["ingredient_names"]:
                    ing_for_label.append((
                        sub_ing[0],
                        (sub_ing[1] * sub_scale) if sub_ing[1] else 0.0,
                        sub_ing[2] if len(sub_ing) > 2 else False,
                    ))
            else:
                warnings.append(
                    f"⚠️ Sub-receta '{name}' no encontrada — "
                    "alérgenos no calculados para este componente."
                )
        else:
            # Leaf ingredient — get effective allergen values
            if not line["category"] and not line["allergen_override"]:
                warnings.append(
                    f"⚠️ '{name}' sin categoría asignada — "
                    "alérgenos no incluidos."
                )
                continue

            eff = _effective_allergens(line)
            accumulated = _union_allergens(accumulated, eff)

            # Add to label ingredient list
            label_name = (
                line.get("label_name_es")
                or line.get("category_label")
                or name
            )
            if label_name:
                grams        = _to_label_grams(name, amount)
                eff          = _effective_allergens(line)
                has_allergen = any(v > 0 for v in eff.values())
                comp2        = line.get("label_name_es_2")
                pct2         = float(line.get("label_name_es_2_pct") or 0) / 100
                comp3        = line.get("label_name_es_3")
                pct3         = float(line.get("label_name_es_3_pct") or 0) / 100
                total_comp   = pct2 + pct3

                if comp2 and 0 < total_comp < 1.0:
                    pct1 = 1.0 - total_comp
                    # Components inherit parent allergen status (backlog item 10)
                    ing_for_label.append((label_name, grams * pct1, has_allergen))
                    ing_for_label.append((comp2,      grams * pct2, has_allergen))
                    if comp3 and pct3 > 0:
                        ing_for_label.append((comp3, grams * pct3, has_allergen))
                else:
                    ing_for_label.append((label_name, grams, has_allergen))

            # Flag ingredients needing verification
            notes = line.get("allergen_notes") or ""
            if "verificar" in notes.lower() or "needs" in notes.lower():
                warnings.append(
                    f"⚠️ '{name}': {notes}"
                )

    # Build Contiene and Puede contener lists
    contiene       = []
    puede_contener = []

    for field in ALLERGEN_FIELDS:
        val = accumulated.get(field, 0)
        display = ALLERGEN_DISPLAY_ES[field]
        if val == 1:
            contiene.append(display)
        elif val == 2:
            puede_contener.append(display)

    # Add kitchen-level may_contain
    # Parse as comma-separated list and add any new items
    if kitchen:
        kitchen_items = [
            k.strip() for k in kitchen.split(",")
            if k.strip()
        ]
        for item in kitchen_items:
            # Only add if not already covered by ingredient-level flags
            if item.lower() not in " ".join(puede_contener).lower() \
                    and item.lower() not in " ".join(contiene).lower():
                puede_contener.append(item)

    return {
        "contiene":         contiene,
        "puede_contener":   puede_contener,
        "accumulated":      accumulated,
        "warnings":         warnings,
        "ingredient_names": ing_for_label,
    }


def get_ingredient_label_text(recipe_id: str) -> dict:
    """
    Generate a draft ingredient label text for a recipe ficha.

    Returns:
    {
        "label_text":  "Harina de trigo, azúcar, mantequilla, huevo, ...",
        "ordered":     [(label_name_es, total_amount_g), ...] sorted desc,
        "warnings":    [list of warning strings],
        "allergen_fields": {field: label_name_es} for bolding in PDF,
    }

    The label text lists ingredients ordered by weight descending (EU legal
    requirement). Allergen-containing ingredients are noted for bolding.
    """
    declaration = get_allergen_declaration(recipe_id)
    ing_names   = declaration["ingredient_names"]
    warnings    = list(declaration["warnings"])

    if not ing_names:
        return {
            "label_text":     "",
            "ordered":        [],
            "warnings":       warnings + ["Sin ingredientes encontrados."],
            "allergen_fields": {},
        }

    # Aggregate by label name — same label from different ingredients
    # (e.g. egg yolk + egg white both become "huevo")
    # ing_names tuples are (label, amount_g, has_allergen)
    aggregated:        dict[str, float] = {}
    allergen_by_label: dict[str, bool]  = {}
    for entry in ing_names:
        label        = entry[0]
        amount       = entry[1]
        has_allergen = entry[2] if len(entry) > 2 else False
        aggregated[label]        = aggregated.get(label, 0) + (amount or 0)
        allergen_by_label[label] = allergen_by_label.get(label, False) or has_allergen

    # Sort by amount descending
    ordered = sorted(aggregated.items(), key=lambda x: x[1], reverse=True)

    # Build the draft text — comma separated, capitalise first word
    ingredient_list = ", ".join(name for name, _ in ordered)
    if ingredient_list:
        ingredient_list = ingredient_list[0].upper() + ingredient_list[1:]

    # Build allergen_fields map directly from ingredient data —
    # {label_name_es: True} for any label that carries an allergen.
    # Used by apply_allergen_bold() for PDF bold rendering.
    allergen_labels = {
        label: True
        for label, has_alg in allergen_by_label.items()
        if has_alg
    }

    return {
        "label_text":      ingredient_list,
        "ordered":         ordered,
        "warnings":        warnings,
        "allergen_fields": allergen_labels,
    }


def apply_allergen_bold(label_text: str, allergen_labels: dict) -> str:
    """
    Wrap allergen-bearing ingredient names in **markers** for bold rendering.
    allergen_labels is the dict returned by get_ingredient_label_text()
    under the key "allergen_fields": {label_name_es: [allergen_fields]}.

    Only the first occurrence of each allergen name is wrapped (per EU 1169/2011
    which requires allergens to be emphasised but not necessarily every instance).

    Returns the label text with allergen names wrapped in ** for bold.
    Capitalisation of the first letter of the whole string is preserved.
    """
    if not label_text or not allergen_labels:
        return label_text

    result = label_text
    for name in sorted(allergen_labels.keys(), key=len, reverse=True):
        # Match case-insensitively, whole word/phrase
        pattern = re.compile(re.escape(name), re.IGNORECASE)
        match   = pattern.search(result)
        if match:
            result = result[:match.start()] + f"**{match.group()}**" + result[match.end():]

    return result


# -----------------------------------------------------------------------------
# Reference data (cake codes, size tiers, price channels)
# These rarely change so we cache them for the session.
# -----------------------------------------------------------------------------

@st.cache_data(ttl=3600)
def get_cake_codes() -> list[dict]:
    sb = get_client()
    result = sb.table("cake_codes").select("*").order("code").execute()
    return result.data or []


@st.cache_data(ttl=3600)
def get_size_tiers() -> list[dict]:
    sb = get_client()
    result = sb.table("size_tiers").select("*").order("code").execute()
    return result.data or []


@st.cache_data(ttl=3600)
def get_price_channels() -> list[dict]:
    sb = get_client()
    result = sb.table("price_channels").select("*").order("code").execute()
    return result.data or []

def save_cake_code(code: str, name: str) -> dict:
    """
    Insert a new cake code. Returns the created row.
    Raises if code already exists (unique constraint).
    """
    sb     = get_client()
    result = (sb.table("cake_codes")
                .insert({"code": code.strip().upper(),
                         "name": name.strip()})
                .execute())
    return result.data[0] if result.data else {}
 

# -----------------------------------------------------------------------------
# SKUs
# -----------------------------------------------------------------------------

def get_skus() -> list[dict]:
    sb = get_client()
    result = (
        sb.table("skus")
        .select("*, recipes(name), size_tiers(code, label), price_channels(code, label)")
        .order("sku_code")
        .execute()
    )
    return result.data or []


def save_sku(record: dict) -> dict:
    sb = get_client()
    if record.get("id"):
        sb.table("skus").update(record).eq("id", record["id"]).execute()
        result = sb.table("skus").select("*").eq("id", record["id"]).execute()
    else:
        result = sb.table("skus").insert(record).execute()
    return result.data[0] if result.data else {}

def get_sku_to_recipe_map() -> list[dict]:
    """
    Build a complete SKU → recipe_id map using two sources:
 
    Source A (primary): holded_products table
      SKU format: CC-01-SS-CH (e.g. LP-01-LA-GW)
      CC = cake code → look up recipe via cake_codes table
      Covers all 90 products in the inventory without manual variant entry.
 
    Source B (override): product_variants table (sku_ws / sku_gw)
      Explicit per-variant SKUs entered via the variants screen.
      Overrides Source A for any SKU that has been manually defined.
 
    Format matches get_skus() so _build_sku_map() in screen_kpis.py
    works unchanged.
    """
    sb = get_client()
 
    # ── Build cake_code → recipe_id lookup ────────────────────────────────────
    # Use the latest version of each recipe per cake code
    cake_codes = get_cake_codes()
    code_by_id = {cc["id"]: cc["code"] for cc in cake_codes}
 
    recipes = (sb.table("recipes")
                 .select("id, cake_code_id, version")
                 .eq("is_sub_recipe", False)
                 .order("version", desc=True)
                 .execute()
                 .data or [])
 
    # cake_code_str → recipe_id  (first encountered = latest version)
    code_to_recipe: dict[str, str] = {}
    for r in recipes:
        if r.get("cake_code_id"):
            code = code_by_id.get(r["cake_code_id"])
            if code and code not in code_to_recipe:
                code_to_recipe[code] = r["id"]
 
    # ── Source A: holded_products ──────────────────────────────────────────────
    rows: dict[str, dict] = {}   # sku_code → row
 
    products = (sb.table("holded_products")
                  .select("sku")
                  .eq("active", True)
                  .execute()
                  .data or [])
 
    for p in products:
        sku   = (p.get("sku") or "").strip()
        parts = sku.split("-")
        if len(parts) < 2:
            continue
        cake_code = parts[0].upper()
        recipe_id = code_to_recipe.get(cake_code)
        if recipe_id:
            rows[sku] = {"sku_code": sku, "recipe_id": recipe_id}
 
    # ── Source B: product_variants (overrides Source A) ────────────────────────
    variants = (sb.table("product_variants")
                  .select("recipe_id, sku_ws, sku_gw")
                  .execute()
                  .data or [])
 
    for v in variants:
        for field in ("sku_ws", "sku_gw"):
            sku = (v.get(field) or "").strip()
            if sku and v.get("recipe_id"):
                rows[sku] = {"sku_code": sku, "recipe_id": v["recipe_id"]}
 
    return list(rows.values())
 


# -----------------------------------------------------------------------------
# Packaging presets
# -----------------------------------------------------------------------------

def get_packaging_presets() -> list[dict]:
    sb = get_client()
    result = sb.table("packaging_presets").select("*").order("name").execute()
    return result.data or []


def get_preset_lines(preset_id: str) -> list[dict]:
    sb = get_client()
    result = (
        sb.table("packaging_preset_lines")
        .select("*, consumables(name, cost_per_unit)")
        .eq("preset_id", preset_id)
        .execute()
    )
    lines = []
    for row in result.data or []:
        con = row.pop("consumables", None) or {}
        row["consumable_name"]      = con.get("name", "")
        row["consumable_cost_per_unit"] = con.get("cost_per_unit")
        lines.append(row)
    return lines


def save_preset(name: str, lines: list[dict], 
                units_per_pack: int = 1) -> None:
    sb = get_client()
    result = sb.table("packaging_presets").insert({
        "name": name,
        "units_per_pack": units_per_pack
    }).execute()
    if not result.data:
        return
    preset_id = result.data[0]["id"]
    for line in lines:
        line["preset_id"] = preset_id
    if lines:
        sb.table("packaging_preset_lines").insert(lines).execute()


def update_preset(preset_id: str, name: str, lines: list[dict],
                  units_per_pack: int = 1) -> None:
    sb = get_client()
    sb.table("packaging_presets").update({
        "name": name,
        "units_per_pack": units_per_pack
    }).eq("id", preset_id).execute()
    sb.table("packaging_preset_lines").delete().eq(
        "preset_id", preset_id
    ).execute()
    if lines:
        for line in lines:
            line["preset_id"] = preset_id
        sb.table("packaging_preset_lines").insert(lines).execute()


def delete_preset(preset_id: str) -> None:
    sb = get_client()
    sb.table("packaging_presets").delete().eq("id", preset_id).execute()

# -----------------------------------------------------------------------------
# Settings
# -----------------------------------------------------------------------------

def get_settings() -> dict:
    sb = get_client()
    result = sb.table("settings").select("*").limit(1).execute()
    return result.data[0] if result.data else {}


def save_settings(record: dict) -> dict:
    sb = get_client()
    if record.get("id"):
        sb.table("settings").update(record).eq("id", record["id"]).execute()
        result = sb.table("settings").select("*").eq("id", record["id"]).execute()
    else:
        result = sb.table("settings").insert(record).execute()
    return result.data[0] if result.data else {}


# =============================================================================
# Price approval and client pricing
# =============================================================================

def get_all_variants_full_with_approval() -> list[dict]:
    """Fetch all variants with working and approved price fields."""
    sb = get_client()
    result = (
        sb.table("product_variants")
        .select(
            "id, recipe_id, format, channel, "
            "ws_price_ex_vat, ws_price_approved, ws_price_approved_at, "
            "rt_price_inc_vat, rt_price_approved, rt_price_approved_at, "
            "ws_price_updated_at, rt_price_updated_at, "
            "size_description"
        )
        .execute()
    )
    return result.data or []


def approve_variant_prices(
    variant_id: str,
    ws_price: float | None,
    rt_price: float | None
) -> None:
    """Copy working prices to approved fields with current timestamp."""
    sb     = get_client()
    record = {"id": variant_id, "updated_at": "now()"}
    if ws_price is not None:
        record["ws_price_approved"]    = ws_price
        record["ws_price_approved_at"] = "now()"
    if rt_price is not None:
        record["rt_price_approved"]    = rt_price
        record["rt_price_approved_at"] = "now()"
    sb.table("product_variants").update(record).eq(
        "id", variant_id
    ).execute()


def get_client_prices() -> list[dict]:
    """Fetch all client-specific prices with variant and recipe info."""
    sb = get_client()
    result = (
        sb.table("client_prices")
        .select(
            "*, "
            "product_variants(format, size_description, "
            "recipes(name))"
        )
        .order("client_name")
        .execute()
    )
    rows = []
    for row in result.data or []:
        variant = row.pop("product_variants", None) or {}
        recipe  = variant.pop("recipes", None) or {}
        fmt     = variant.get("format", "")
        fmt_label = {"standard": "Estándar", "individual": "Individual",
                     "bocado": "Bocado"}.get(fmt, fmt)
        row["variant_label"] = (
            f"{recipe.get('name', '')} — {fmt_label}"
        )
        rows.append(row)
    return rows


def save_client_price(record: dict) -> dict:
    """Save a client-specific price (upsert on client_name + variant_id)."""
    sb = get_client()
    # Check if exists
    existing = (
        sb.table("client_prices")
        .select("id")
        .eq("client_name", record["client_name"])
        .eq("variant_id",  record["variant_id"])
        .execute()
    )
    if existing.data:
        record["id"] = existing.data[0]["id"]
        sb.table("client_prices").update(record).eq(
            "id", record["id"]
        ).execute()
        result = sb.table("client_prices").select("*").eq(
            "id", record["id"]
        ).execute()
    else:
        result = sb.table("client_prices").insert(record).execute()
    return result.data[0] if result.data else {}


def delete_client_price(price_id: str) -> None:
    sb = get_client()
    sb.table("client_prices").delete().eq("id", price_id).execute()


def get_client_prices_for_catalogue(client_name: str) -> dict[str, dict]:
    """
    Return client-specific prices for a named client, keyed by variant_id.
    Used in catalogue generation to override standard approved prices.
    """
    sb = get_client()
    today = str(date.today())
    result = (
        sb.table("client_prices")
        .select("variant_id, ws_price_ex_vat, rt_price_inc_vat")
        .eq("client_name", client_name)
        .lte("valid_from", today)
        .or_(f"valid_until.is.null,valid_until.gte.{today}")
        .execute()
    )
    return {r["variant_id"]: r for r in (result.data or [])}

# -----------------------------------------------------------------------------
# Holded year cache
# Persistent cache for Holded invoice data. One row per calendar year.
# Historical years are written once and never updated automatically.
# -----------------------------------------------------------------------------

def get_holded_cache_index() -> list[dict]:
    """
    Return one summary row per cached year:
      { year, invoice_count, synced_at }
    Used by holded_api.py to know which years are already stored.
    """
    sb     = get_client()
    result = (
        sb.table("holded_year_cache")
        .select("year, synced_at")
        .order("year")
        .execute()
    )
    return result.data or []


def get_holded_year_cache(year: int) -> list[dict]:
    """
    Return the cached list of invoice dicts for a given year.
    Returns [] if the year is not cached.
    """
    sb     = get_client()
    result = (
        sb.table("holded_year_cache")
        .select("invoices")
        .eq("year", year)
        .limit(1)
        .execute()
    )
    if not result.data:
        return []
    return result.data[0].get("invoices") or []


def save_holded_year_cache(year: int, invoices: list[dict],
                           cache_version: int = 1) -> None:
    """
    Upsert invoice data for a given year into the Supabase cache.
    Called once per historical year; never called for the current year.
    """
    sb = get_client()
    sb.table("holded_year_cache").upsert({
        "year":      year,
        "invoices":  invoices,
        "synced_at": "now()",
    }).execute()

# =============================================================================
# Excel parsing helpers
# =============================================================================
 
def _parse_month_header(cell_value: str) -> tuple[int, int] | None:
    """
    Parse a Holded month column header like 'Enero 26' or 'Febrero 2026'
    into (month_number, year). Returns None if not parseable.
    """
    if not cell_value:
        return None
    parts = str(cell_value).lower().split()
    if len(parts) < 2:
        return None
    month_name = parts[0]
    month = _MONTHS_ES.get(month_name)
    if not month:
        return None
    try:
        year_str = parts[1]
        year = int(year_str) if len(year_str) == 4 else 2000 + int(year_str)
        return month, year
    except (ValueError, IndexError):
        return None
 
 
def _extract_sku(product_name: str) -> tuple[str, str | None]:
    """
    If the product name contains an embedded SKU (e.g. 'Cookie Box - CO-03-DC-GW'),
    return (clean_name, sku). Otherwise return (product_name, None).
    """
    m = _SKU_RE.search(product_name)
    if not m:
        return product_name.strip(), None
    sku = m.group(1)
    # Remove the SKU and any trailing separator from the name
    clean = re.sub(r'\s*[-–]\s*' + re.escape(sku) + r'\s*$', '', product_name).strip()
    return clean, sku
 
 
def parse_ventas_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse a Holded 'Ventas' Excel export (monthly revenue totals).
 
    Returns a list of dicts, one per month with data:
      { year, month, ventas_ex_vat, tax, total_inc_vat, units }
 
    Months with ventas_ex_vat == 0 are skipped (future months in the export).
    """
    import openpyxl
    from io import BytesIO
 
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
 
    # Build column index → (month, year) map from the header row
    col_map: dict[int, tuple[int, int]] = {}
    header_row_idx = None
 
    rows = list(ws.iter_rows(values_only=True))
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            parsed = _parse_month_header(str(cell) if cell else "")
            if parsed:
                col_map[col_idx] = parsed
                header_row_idx = row_idx
        if col_map:
            break
 
    if not col_map:
        raise ValueError("No se encontró la fila de cabecera de meses en el fichero de Ventas.")
 
    # Extract metric rows
    metric_map = {
        'ventas':    'ventas_ex_vat',
        'impuestos': 'tax',
        'total':     'total_inc_vat',
        'unidades':  'units',
    }
 
    # {(year, month): {metric: value}}
    data: dict[tuple, dict] = {}
 
    for row in rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip().lower()
        field = metric_map.get(label)
        if not field:
            continue
        for col_idx, (month, year) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            try:
                val = float(val or 0)
            except (TypeError, ValueError):
                val = 0.0
            key = (year, month)
            if key not in data:
                data[key] = {'year': year, 'month': month,
                             'ventas_ex_vat': 0.0, 'tax': 0.0,
                             'total_inc_vat': 0.0, 'units': 0.0}
            data[key][field] = val
 
    # Only return months that have actual revenue (skip future zero months)
    today = date.today()
    return [
        v for v in data.values()
        if v['ventas_ex_vat'] != 0 or v['total_inc_vat'] != 0
        if not (v['year'] == today.year and v['month'] == today.month)
    ]
 
def parse_productos_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse a Holded 'Ventas por producto' Excel export (units per product per month).
 
    Returns a list of dicts:
      { year, month, product_name, sku (or None), units }
 
    Rows with zero units across all months are skipped.
    'Total' row is skipped.
    """
    import openpyxl
    from io import BytesIO
 
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
 
    rows = list(ws.iter_rows(values_only=True))
 
    # Find header row
    col_map: dict[int, tuple[int, int]] = {}
    header_row_idx = None
 
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            parsed = _parse_month_header(str(cell) if cell else "")
            if parsed:
                col_map[col_idx] = parsed
                header_row_idx = row_idx
        if col_map:
            break
 
    if not col_map:
        raise ValueError("No se encontró la fila de cabecera de meses en el fichero de Productos.")
 
    results = []
    skip_labels = {'total', 'informe creado'}
 
    for row in rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue
        raw_name = str(row[0]).strip()
        if not raw_name:
            continue
        if any(raw_name.lower().startswith(s) for s in skip_labels):
            continue
 
        clean_name, sku = _extract_sku(raw_name)
 
        for col_idx, (month, year) in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            try:
                units = float(val or 0)
            except (TypeError, ValueError):
                units = 0.0
            if units == 0:
                continue
            results.append({
                'year':         year,
                'month':        month,
                'product_name': clean_name,
                'sku':          sku,
                'units':        units,
            })
 
    today = date.today()
    
    # Deduplicate by (year, month, product_name) — sum units if duplicate
    deduped: dict[tuple, dict] = {}
    for r in results:
        if r['year'] == today.year and r['month'] == today.month:
            continue
        key = (r['year'], r['month'], r['product_name'])
        if key in deduped:
            deduped[key]['units'] += r['units']
        else:
            deduped[key] = r
    return list(deduped.values())

 
# =============================================================================
# Supabase read/write — monthly revenue
# =============================================================================
 
def upsert_monthly_revenue(rows: list[dict]) -> int:
    """
    Upsert monthly revenue rows into holded_monthly_revenue.
    Returns number of rows upserted.
    """
    if not rows:
        return 0
    sb = get_client()
    payload = [
        {
            'year':          r['year'],
            'month':         r['month'],
            'ventas_ex_vat': r['ventas_ex_vat'],
            'tax':           r['tax'],
            'total_inc_vat': r['total_inc_vat'],
            'units':         r['units'],
            'uploaded_at':   'now()',
        }
        for r in rows
    ]
    sb.table('holded_monthly_revenue').upsert(payload).execute()
    return len(payload)
 
 
def get_monthly_revenue(year: int | None = None) -> list[dict]:
    """
    Return all monthly revenue rows, optionally filtered by year.
    Sorted by year, month ascending.
    """
    sb = get_client()
    q  = sb.table('holded_monthly_revenue').select('*').order('year').order('month')
    if year is not None:
        q = q.eq('year', year)
    return q.execute().data or []
 
 
def upsert_monthly_products(rows: list[dict]) -> int:
    """
    Upsert monthly product rows into holded_monthly_products.
    Returns number of rows upserted.
    """
    if not rows:
        return 0
    sb = get_client()
    payload = [
        {
            'year':         r['year'],
            'month':        r['month'],
            'product_name': r['product_name'],
            'sku':          r.get('sku'),
            'units':        r['units'],
            'uploaded_at':  'now()',
        }
        for r in rows
    ]
    sb.table('holded_monthly_products').upsert(payload).execute()
    return len(payload)
 
 
def get_monthly_products(year: int | None = None,
                         month: int | None = None) -> list[dict]:
    """
    Return product rows, optionally filtered by year and/or month.
    """
    sb = get_client()
    q  = (sb.table('holded_monthly_products')
            .select('*')
            .order('year').order('month').order('units', desc=True))
    if year  is not None: q = q.eq('year',  year)
    if month is not None: q = q.eq('month', month)
    return q.execute().data or []
 
 
def get_upload_status() -> dict:
    """
    Return a summary of what data has been uploaded:
      {
        'months':      [(year, month), ...],   # all uploaded months
        'latest_year': int | None,
        'latest_month': int | None,
      }
    """
    sb   = get_client()
    rows = (sb.table('holded_monthly_revenue')
              .select('year, month, uploaded_at')
              .order('year', desc=True).order('month', desc=True)
              .execute().data or [])
    months = [(r['year'], r['month']) for r in rows]
    return {
        'months':       months,
        'latest_year':  rows[0]['year']  if rows else None,
        'latest_month': rows[0]['month'] if rows else None,
        'latest_upload': rows[0]['uploaded_at'] if rows else None,
    }
 
    
# =============================================================================
# holded_products — inventory / product catalogue
# =============================================================================
 
def parse_inventory_excel(file_bytes: bytes) -> list[dict]:
    import openpyxl
    from io import BytesIO

    wb   = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().upper() == 'SKU':
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("No se encontró la fila de cabecera 'SKU' en el fichero.")

    SKU_RE = re.compile(
        r'^[A-Z]{2}-?\d{2}-?[A-Z]{2}-?[A-Z]{2,4}(?:-[A-Z]{2,4})?$'
    )

    def _normalise_sku(raw: str) -> str | None:
        if not raw:
            return None
        s = raw.strip().upper()
        # Correct common OCR/typo: digit 0 in position 0 or 1 replaced with letter O
        s = s[:2].replace('0', 'O') + s[2:]
        if SKU_RE.match(s):
            return s
        m = re.match(r'^([A-Z]{2})(\d{2})([A-Z]{2})([A-Z]{2,4})(?:([A-Z]{2,4}))?$', s)
        if m:
            parts = [p for p in m.groups() if p]
            return '-'.join(parts)
        return None

    seen: dict[tuple, dict] = {}
    skip_names = {'informe creado', 'none', ''}

    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        raw_sku  = str(row[0]).strip() if row[0] else ''
        raw_name = str(row[1]).strip() if row[1] else ''
        
        if not raw_name or raw_name.lower() in skip_names:
            continue

        sku = _normalise_sku(raw_sku)
        if not sku:
            continue

        try:
            price = float(row[4]) if row[4] and str(row[4]) not in ('-', 'None') else None
        except (TypeError, ValueError):
            price = None

        key = (sku, raw_name)
        if key not in seen or (price and (seen[key]['price_ex_vat'] or 0) < price):
            seen[key] = {
                'sku':          sku,
                'name':         raw_name,
                'price_ex_vat': price,
            }

    return list(seen.values())
 
 
def upsert_holded_products(rows: list[dict]) -> int:
    """Upsert product catalogue rows. Returns count upserted."""
    if not rows:
        return 0
    sb = get_client()
    payload = [
        {
            'sku':          r['sku'],
            'name':         r['name'],
            'price_ex_vat': r.get('price_ex_vat'),
            'uploaded_at':  'now()',
        }
        for r in rows
    ]
    sb.table('holded_products').upsert(payload).execute()
    return len(payload)
 
 
def get_holded_products() -> list[dict]:
    """Return all active products: [{sku, name, price_ex_vat}]."""
    sb = get_client()
    return (
        sb.table('holded_products')
        .select('sku, name, price_ex_vat, units_per_pack')
        .eq('active', True)
        .order('name')
        .execute()
        .data or []
    )
 
 
def get_name_to_sku_map() -> dict[str, str]:
    """
    Return {product_name: sku} for all active products.
 
    When multiple SKUs share the same name (e.g. Roscón large with different
    fillings), the first SKU alphabetically is used and the ambiguous names
    are stored in st.session_state['_holded_ambiguous_names'] so the UI
    can surface a warning.
    """
 
    products   = get_holded_products()
    result:    dict[str, str]       = {}
    ambiguous: dict[str, list[str]] = {}
 
    for p in sorted(products, key=lambda x: x['sku']):
        name = p['name']
        if name in result:
            ambiguous.setdefault(name, [result[name]]).append(p['sku'])
        else:
            result[name] = p['sku']
 
    st.session_state['_holded_ambiguous_names'] = ambiguous
    return result

def _next_lote_number(sb, prod_date: "date") -> str:
    """
    Generate the next sequential lote number for a given date.
    Format: MC-YYMMDD-XXX  (XXX = 001, 002, …)
    Uses a date-range filter instead of LIKE to avoid Cloudflare edge issues
    with % wildcards in query parameters.
    """
    import datetime as _dt
    date_str  = prod_date.strftime("%y%m%d")
    prefix    = f"MC-{date_str}-"
    day_start = _dt.datetime.combine(prod_date, _dt.time.min).isoformat()
    day_end   = _dt.datetime.combine(prod_date, _dt.time.max).isoformat()
    existing  = (
        sb.table("production_runs")
          .select("lote_number")
          .gte("production_date", day_start)
          .lte("production_date", day_end)
          .execute()
          .data or []
    )
    # Count only rows that actually have the expected prefix (guard against
    # multiple runs on the same date from different sources)
    count = sum(1 for r in existing if (r.get("lote_number") or "").startswith(prefix))
    return f"{prefix}{count + 1:03d}"
 
 
def save_production_run(
    recipe_id:         str,
    recipe_name:       str,
    fmt:               str,
    prod_date:         "date",
    quantity:          float,
    quantity_unit:     str = "units",
    oven_temp_c:       float | None = None,
    bake_time_min:     int | None = None,
    notes:             str | None = None,
    ing_refs:          list[dict] | None = None,
    pcc_log:           list[dict] | None = None,
    is_component_run:  bool = False,
    amount_produced_g: float | None = None,
) -> dict:
    """
    Insert a production run + ingredient refs.
    Returns the saved run dict (with lote_number and id).
    """
    import json as _json
    if ing_refs is None:
        ing_refs = []
    sb   = get_client()
    lote = _next_lote_number(sb, prod_date)

    run_row = {
        "lote_number":     lote,
        "recipe_id":       recipe_id,
        "recipe_name":     recipe_name,
        "format":          fmt,
        "production_date": prod_date.isoformat(),
        "quantity":        quantity,
        "quantity_unit":   quantity_unit,
        "oven_temp_c":     oven_temp_c,
        "bake_time_min":   bake_time_min,
        "notes":           notes,
        "pcc_log":         _json.dumps(pcc_log) if pcc_log else None,
        "is_component_run":  is_component_run,
        "amount_produced_g": amount_produced_g,
    }
    result = (
        sb.table("production_runs")
          .insert(run_row)
          .execute()
    )
    saved = (result.data or [{}])[0]
    run_id = saved["id"]
 
    # Insert ingredient refs
    for ref in ing_refs:
        sb.table("production_ingredient_refs").insert({
            "production_run_id": run_id,
            "ingredient_name":   ref.get("ingredient_name"),
            "albaran_ref":       ref.get("albaran_ref"),
        }).execute()
 
    saved["ingredient_refs"] = ing_refs
    return saved
 
 
def get_production_runs(limit: int = 30) -> list[dict]:
    """Return recent production runs, newest first."""
    sb   = get_client()
    rows = (
        sb.table("production_runs")
          .select("*")
          .order("production_date", desc=True)
          .order("lote_number",     desc=True)
          .limit(limit)
          .execute()
          .data or []
    )
    # Attach ingredient refs to each run
    if rows:
        run_ids = [r["id"] for r in rows]
        refs    = (
            sb.table("production_ingredient_refs")
              .select("*")
              .in_("production_run_id", run_ids)
              .execute()
              .data or []
        )
        refs_by_run: dict[str, list] = {}
        for ref in refs:
            refs_by_run.setdefault(ref["production_run_id"], []).append(ref)
        for row in rows:
            row["ingredient_refs"] = refs_by_run.get(row["id"], [])
    return rows
 
 
def get_production_run(run_id: str) -> dict | None:
    """Return a single production run with ingredient refs."""
    sb   = get_client()
    rows = (
        sb.table("production_runs")
          .select("*")
          .eq("id", run_id)
          .limit(1)
          .execute()
          .data or []
    )
    if not rows:
        return None
    run  = rows[0]
    refs = (
        sb.table("production_ingredient_refs")
          .select("*")
          .eq("production_run_id", run_id)
          .execute()
          .data or []
    )
    run["ingredient_refs"] = refs
    return run

 
def get_key_ingredients_for_recipe(recipe_id: str) -> list[dict]:
    """
    Return the key purchased ingredients for a recipe, recursively resolving
    sub-recipes using the same expansion logic as get_allergen_declaration.
 
    "Key" means:
      - >=5% of total final product weight, OR
      - bears any declared allergen (value > 0 on any ALLERGEN_FIELDS)
        determined from the ingredient's actual allergen profile
        (ingredient override or category), consistent with the ficha.
 
    Returns list of {name, quantity_g, pct, is_allergen_bearing} dicts,
    ordered by quantity_g descending.
    """
    def _resolve(rid: str, scale: float, depth: int,
                 visited: set) -> list[tuple]:
        """
        Recursively expand recipe into flat list of
        (ingredient_name, grams_at_scale, has_allergen) tuples.
        Uses _to_label_grams (module-level) for unit conversion — no duplication.
        """
        if depth > 5 or rid in visited:
            return []
        visited = visited | {rid}

        lines  = _get_recipe_lines_with_allergens(rid)
        result = []

        for line in lines:
            name       = line["ingredient_name"]
            raw_amount = line["amount"] * scale
            amount     = _to_label_grams(name, raw_amount)

            if line["is_sub_recipe"]:
                comp_id = line.get("component_recipe_id")
                sub     = get_recipe(comp_id) if comp_id else _find_recipe_by_ingredient_name(name)
                if not sub:
                    continue
                sub_lines   = _get_recipe_lines_with_allergens(sub["id"])
                sub_total_g = sum(
                    _to_label_grams(l["ingredient_name"], l["amount"])
                    for l in sub_lines
                ) or amount
                sub_scale   = amount / sub_total_g
                result.extend(_resolve(sub["id"], sub_scale, depth + 1, visited))
            else:
                eff          = _effective_allergens(line)
                has_allergen = any(v > 0 for v in eff.values())
                # Split compound ingredients by component percentages
                comp2 = line.get("label_name_es_2")
                pct2  = float(line.get("label_name_es_2_pct") or 0) / 100
                comp3 = line.get("label_name_es_3")
                pct3  = float(line.get("label_name_es_3_pct") or 0) / 100
                total_comp = pct2 + pct3
                if comp2 and 0 < total_comp < 1.0:
                    pct1 = 1.0 - total_comp
                    result.append((line.get("label_name_es") or name, amount * pct1, has_allergen))
                    result.append((comp2, amount * pct2, has_allergen))
                    if comp3 and pct3 > 0:
                        result.append((comp3, amount * pct3, has_allergen))
                else:
                    result.append((name, amount, has_allergen))

        return result
 
    flat = _resolve(recipe_id, 1.0, 0, set())
    if not flat:
        return []
 
    # Aggregate by ingredient name (same ingredient may appear in multiple sub-recipes)
    totals:    dict[str, float] = {}
    allergens: dict[str, bool]  = {}
    for name, grams, has_alg in flat:
        totals[name]    = totals.get(name, 0) + grams
        allergens[name] = allergens.get(name, False) or has_alg
 
    total_g = sum(totals.values())
 
    result = []
    for name, g in totals.items():
        pct         = (g / total_g * 100) if total_g else 0
        is_allergen = allergens[name]
        if pct >= 5.0 or is_allergen:
            result.append({
                "name":                name,
                "quantity_g":          round(g, 1),
                "pct":                 round(pct, 1),
                "is_allergen_bearing": is_allergen,
            })
 
    result.sort(key=lambda x: x["quantity_g"], reverse=True)
    return result
 
 
def get_production_runs_for_recipe(recipe_id: str, limit: int = 1) -> list[dict]:
    """Return the most recent production runs for a specific recipe (with ingredient refs)."""
    sb   = get_client()
    rows = (
        sb.table("production_runs")
          .select("*")
          .eq("recipe_id", recipe_id)
          .order("production_date", desc=True)
          .order("created_at",      desc=True)
          .limit(limit)
          .execute()
          .data or []
    )
    if rows:
        run_ids = [r["id"] for r in rows]
        refs    = (
            sb.table("production_ingredient_refs")
              .select("*")
              .in_("production_run_id", run_ids)
              .execute()
              .data or []
        )
        refs_by_run: dict[str, list] = {}
        for ref in refs:
            refs_by_run.setdefault(ref["production_run_id"], []).append(ref)
        for row in rows:
            row["ingredient_refs"] = refs_by_run.get(row["id"], [])
    return rows
 
def update_production_run(run_id: str, updates: dict) -> None:
    """Update editable fields on an existing production run."""
    import json as _json
    if "pcc_log" in updates and not isinstance(updates["pcc_log"], str):
        updates["pcc_log"] = _json.dumps(updates["pcc_log"])
    get_client().table("production_runs").update(updates).eq("id", run_id).execute()


def replace_production_ingredient_refs(run_id: str, refs: list[dict]) -> None:
    """Delete and re-insert ingredient refs for a production run."""
    sb = get_client()
    sb.table("production_ingredient_refs").delete().eq("production_run_id", run_id).execute()
    for ref in refs:
        if ref.get("ingredient_name"):
            sb.table("production_ingredient_refs").insert({
                "production_run_id": run_id,
                "ingredient_name":   ref["ingredient_name"],
                "albaran_ref":       ref.get("albaran_ref"),
            }).execute()


def delete_production_run(run_id: str) -> None:
    """Delete a single production run and its ingredient refs."""
    sb = get_client()
    sb.table("production_ingredient_refs").delete().eq("production_run_id", run_id).execute()
    sb.table("production_runs").delete().eq("id", run_id).execute()


def update_goods_receipt(receipt_id: str, header: dict, items: list[dict]) -> None:
    """Update a goods receipt header and replace its items."""
    sb = get_client()
    sb.table("goods_receipts").update(header).eq("id", receipt_id).execute()
    sb.table("goods_receipt_items").delete().eq("receipt_id", receipt_id).execute()
    for item in items:
        row = {k: v for k, v in item.items() if k not in ("id", "receipt_id")}
        row["receipt_id"] = receipt_id
        sb.table("goods_receipt_items").insert(row).execute()


def delete_goods_receipt(receipt_id: str) -> None:
    """Delete a single goods receipt and its items."""
    sb = get_client()
    sb.table("goods_receipt_items").delete().eq("receipt_id", receipt_id).execute()
    sb.table("goods_receipts").delete().eq("id", receipt_id).execute()


def delete_production_runs_before(cutoff_date) -> int:
    """
    Delete all production runs (and their ingredient refs) with
    production_date strictly before cutoff_date. Returns the count deleted.
    """
    sb = get_client()
    iso = cutoff_date.isoformat() if hasattr(cutoff_date, "isoformat") else str(cutoff_date)
    # Fetch IDs first so we can cascade-delete refs
    run_ids = [
        r["id"] for r in (
            sb.table("production_runs").select("id").lt("production_date", iso).execute().data or []
        )
    ]
    if not run_ids:
        return 0
    sb.table("production_ingredient_refs").delete().in_("production_run_id", run_ids).execute()
    sb.table("production_runs").delete().lt("production_date", iso).execute()
    return len(run_ids)


# -----------------------------------------------------------------------------
# Component production runs
# -----------------------------------------------------------------------------

def save_component_production_run(
    recipe_id:         str,
    recipe_name:       str,
    prod_date:         "date",
    amount_produced_g: float,
    notes:             str | None = None,
    ing_refs:          list[dict] | None = None,
    pcc_log:           list[dict] | None = None,
) -> dict:
    """Save a component recipe production run."""
    return save_production_run(
        recipe_id         = recipe_id,
        recipe_name       = recipe_name,
        fmt               = "component",
        prod_date         = prod_date,
        quantity          = amount_produced_g,
        quantity_unit     = "g",
        notes             = notes,
        ing_refs          = ing_refs or [],
        pcc_log           = pcc_log,
        is_component_run  = True,
        amount_produced_g = amount_produced_g,
    )


def get_component_production_runs(limit: int = 50) -> list[dict]:
    """Return recent component production runs for use in pickers."""
    sb = get_client()
    return (
        sb.table("production_runs")
          .select("id, lote_number, recipe_id, recipe_name, production_date, amount_produced_g")
          .eq("is_component_run", True)
          .order("production_date", desc=True)
          .order("created_at",     desc=True)
          .limit(limit)
          .execute()
          .data or []
    )


def link_component_runs(final_run_id: str, component_run_ids: list[str]) -> None:
    """Replace the set of component production runs linked to a final recipe run."""
    sb = get_client()
    sb.table("production_run_component_refs").delete().eq("production_run_id", final_run_id).execute()
    for comp_run_id in component_run_ids:
        sb.table("production_run_component_refs").insert({
            "production_run_id": final_run_id,
            "component_run_id":  comp_run_id,
        }).execute()


def get_component_refs_for_run(run_id: str) -> list[dict]:
    """Return component production runs linked to a final recipe run."""
    sb = get_client()
    refs = (
        sb.table("production_run_component_refs")
          .select(
              "id, component_run_id, "
              "component_run:component_run_id(lote_number, recipe_name, production_date, amount_produced_g)"
          )
          .eq("production_run_id", run_id)
          .execute()
          .data or []
    )
    return refs


def calc_component_cost_per_g(recipe_id: str, labour_rate_per_hour: float = 0.0) -> float:
    """
    Compute ingredient-only cost per gram for a component recipe.

    Returns ingredient_cost / ref_weight_g.  Labour (labour_per_kg) is
    intentionally excluded here so that the calling analysis screen can
    account for it separately in the labour bucket rather than the
    ingredients bucket.  The labour_rate_per_hour parameter is accepted
    for backward compatibility but is no longer used.

    Uses get_ingredients() for cost lookup so derived ingredient costs
    (clara, yema, zumo) are resolved correctly rather than reading the raw
    null cost_per_unit from the DB.
    """
    recipe       = get_recipe(recipe_id)
    ref_weight_g = float(recipe.get("ref_weight_kg") or 1) * 1000

    # Use get_ingredients() so derived-ingredient costs are computed
    all_ings = get_ingredients()
    ing_map  = {i["name"]: i for i in all_ings}

    lines           = get_recipe_lines(recipe_id)
    ingredient_cost = 0.0
    for line in lines:
        if line.get("is_component_line"):
            continue  # nested components not supported (two-level rule)
        name   = line.get("ingredient_name", "")
        cpu    = (ing_map.get(name) or {}).get("cost_per_unit")
        amount = float(line.get("amount") or 0)
        if cpu:
            ingredient_cost += float(cpu) * amount

    return ingredient_cost / ref_weight_g if ref_weight_g > 0 else 0.0


# -----------------------------------------------------------------------------
# Variant migration
# -----------------------------------------------------------------------------

def reassign_variants(from_recipe_id: str, to_recipe_id: str) -> int:
    """
    Move all variants from from_recipe_id to to_recipe_id.
    Resets label_approved=False on all moved variants (allergen re-review required).
    Returns count of variants reassigned.
    """
    sb       = get_client()
    variants = (
        sb.table("product_variants")
          .select("id")
          .eq("recipe_id", from_recipe_id)
          .execute()
          .data or []
    )
    if not variants:
        return 0
    ids = [v["id"] for v in variants]
    (sb.table("product_variants")
       .update({"recipe_id": to_recipe_id, "label_approved": False, "updated_at": "now()"})
       .in_("id", ids)
       .execute())
    return len(ids)


# PCC STEPS

def get_pcc_steps(recipe_id: str) -> list[dict]:
    """Return all PCC steps for a recipe, ordered by sort_order."""
    sb = get_client()
    return (
        sb.table("recipe_pcc_steps")
          .select("*")
          .eq("recipe_id", recipe_id)
          .order("sort_order")
          .execute()
          .data or []
    )
 
 
def replace_pcc_steps(recipe_id: str, steps: list[dict]) -> None:
    """
    Replace all PCC steps for a recipe.
    Deletes steps that are no longer present, upserts the rest.
    """
    sb = get_client()
 
    # Fetch existing step IDs
    existing = (
        sb.table("recipe_pcc_steps")
          .select("id")
          .eq("recipe_id", recipe_id)
          .execute()
          .data or []
    )
    existing_ids = {r["id"] for r in existing}
    incoming_ids = {s["id"] for s in steps if s.get("id")}
 
    # Delete removed steps
    ids_to_delete = existing_ids - incoming_ids
    if ids_to_delete:
        (sb.table("recipe_pcc_steps")
           .delete()
           .in_("id", list(ids_to_delete))
           .execute())

    # Upsert remaining
    for i, step in enumerate(steps):
        row = {
            "recipe_id":             recipe_id,
            "step_name":             step["step_name"],
            "target_temp_c":         step.get("target_temp_c"),
            "target_time_min":       step.get("target_time_min"),
            "critical_limit_temp_c": step.get("critical_limit_temp_c") or 70.0,
            "sort_order":            step.get("sort_order", i),
        }
        if step.get("id"):
            (sb.table("recipe_pcc_steps")
               .update(row)
               .eq("id", step["id"])
               .execute())
        else:
            (sb.table("recipe_pcc_steps")
               .insert(row)
               .execute())

# =============================================================================
# TO DO list DB FUNCTIONS
# =============================================================================

def get_todos() -> list[dict]:
    """Return all todos ordered by priority then due date."""
    sb = get_client()
    return (
        sb.table("todos")
          .select("*")
          .order("due_date", desc=False, nullsfirst=False)
          .execute()
          .data or []
    )


def save_todo(record: dict) -> dict:
    """Insert or update a todo. Returns the saved record."""
    sb = get_client()
    if record.get("id"):
        sb.table("todos").update({
            k: v for k, v in record.items() if k != "id"
        }).eq("id", record["id"]).execute()
        return record
    else:
        result = sb.table("todos").insert(record).execute()
        return (result.data or [{}])[0]


def delete_todo(todo_id: str) -> None:
    """Delete a todo by id."""
    get_client().table("todos").delete().eq("id", todo_id).execute()

# =============================================================================
# GOODS RECEIPTS  (Registro de Recepción — APPCC ELD R7-01)
# =============================================================================
 
def save_goods_receipt(
    *,
    receipt_date: "date",
    supplier: str,
    albaran_ref: str | None,
    received_by: str | None,
    items: list[dict],          # list of item dicts — see schema
    notes: str | None,
) -> dict:
    """
    Save one goods-receipt header + its line items.
    Returns the saved header dict with 'id' populated.
    """
    sb = get_client()
 
    header = {
        "receipt_date":  receipt_date.isoformat(),
        "supplier":      supplier.strip(),
        "albaran_ref":   (albaran_ref or "").strip() or None,
        "received_by":   (received_by or "").strip() or None,
        "notes":         (notes or "").strip() or None,
    }
    result = sb.table("goods_receipts").insert(header).execute()
    saved  = (result.data or [{}])[0]
    rid    = saved["id"]
 
    for item in items:
        sb.table("goods_receipt_items").insert({
            "receipt_id":          rid,
            "product_name":        item.get("product_name", "").strip(),
            "supplier_lot":        (item.get("supplier_lot") or "").strip() or None,
            "quantity":            item.get("quantity"),
            "quantity_unit":       item.get("quantity_unit"),
            "temp_type":           item.get("temp_type", "ambient"),   # 'refrigerated'|'frozen'|'ambient'
            "temp_measured_c":     item.get("temp_measured_c"),        # None for ambient
            "temp_limit_c":        item.get("temp_limit_c"),           # None for ambient
            "packaging_ok":        item.get("packaging_ok", True),
            "labelling_ok":        item.get("labelling_ok", True),
            "accepted":            item.get("accepted", True),
            "rejection_reason":    (item.get("rejection_reason") or "").strip() or None,
        }).execute()
 
    saved["items"] = items
    return saved
 
 
def get_goods_receipts(limit: int = 30) -> list[dict]:
    """Return recent goods receipts, newest first, with items attached."""
    sb   = get_client()
    rows = (
        sb.table("goods_receipts")
          .select("*")
          .order("receipt_date", desc=True)
          .order("created_at",   desc=True)
          .limit(limit)
          .execute()
          .data or []
    )
    if not rows:
        return rows
    receipt_ids = [r["id"] for r in rows]
    items = (
        sb.table("goods_receipt_items")
          .select("*")
          .in_("receipt_id", receipt_ids)
          .execute()
          .data or []
    )
    items_by_receipt: dict[str, list] = {}
    for it in items:
        items_by_receipt.setdefault(it["receipt_id"], []).append(it)
    for row in rows:
        row["items"] = items_by_receipt.get(row["id"], [])
    return rows
